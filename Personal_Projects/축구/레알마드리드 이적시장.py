import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager, rc
from matplotlib.ticker import FuncFormatter
import re
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

# 폰트 설정
font_path = "C:/Windows/Fonts/malgun.TTF"
font = font_manager.FontProperties(fname=font_path).get_name()
rc('font', family=font)

# 데이터 경로
file_paths = ['C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1415.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1516.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1617.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1718.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1819.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_1920.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_2021.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_2122.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_2223.csv',
              'C:/Users/yth21/Desktop/TH/data/Football/datas/RM_transfer_2324.csv']

# 시즌 숫자를 문자열 형태로 리스트에 저장
season_nums = [re.findall(r'\d+', path)[-1] for path in file_paths]
season_lists = [pd.read_csv(file_path, header=2) for file_path in file_paths]

# 시즌 컬럼 추가
for i, season_df in enumerate(season_lists):
    season_df['시즌'] = season_nums[i]

season_count = 0

# 시즌별로 지출과 수입을 저장할 리스트 생성
spending_totals = []
M_spending_totals = []

income_totals = []
M_income_totals = []

# 모든 시즌의 데이터를 하나의 데이터프레임으로
all_season_data = pd.concat(season_lists, ignore_index=True)

# 최대 지출과 최대 수입을 얻은 선수 찾기
max_spending_player = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].min()]['선수 이름'].values[0]
max_spending_player_spend = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].min()]['이적료(유로)'].values[0]
max_spending_player_spend = max_spending_player_spend * -1 / 1e6
max_spending_player_season = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].min()]['시즌'].values[0]

max_income_player = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].max()]['선수 이름'].values[0]
max_income_player_earn = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].max()]['이적료(유로)'].values[0]
max_income_player_earn = max_income_player_earn / 1e6
max_income_player_season = all_season_data[all_season_data['이적료(유로)'] == all_season_data['이적료(유로)'].max()]['시즌'].values[0]

# 10년간 총 지출과 수입, 넷스펜딩 계산
all_spending = all_season_data[all_season_data['이적료(유로)'] < 0]
all_income = all_season_data[all_season_data['이적료(유로)'] > 0]

all_spending = all_spending['이적료(유로)'].sum() * -1 / 1e6
all_income = all_income['이적료(유로)'].sum() / 1e6

all_net_spending = all_spending - all_income

for df in season_lists:
    
    tot_spending = 0
    tot_income = 0
    
    spending_count = 0
    income_count = 0
    
    # 지출과 수입 데이터 추출
    spending_data = df[df['이적 형태'] == 'Arrival']
    income_data = df[df['이적 형태'] == 'Departure']
    
    tot_spending = spending_data['이적료(유로)'].sum()
    tot_spending = tot_spending * -1
    spending_count = len(spending_data)
    
    tot_income = income_data['이적료(유로)'].sum()
    income_count = len(income_data)

    FA_arrival_nums = spending_data[spending_data['기타'] == 'FA'].shape[0]
    Loan_arrival_nums = spending_data[spending_data['기타'] == 'Loan'].shape[0]
    Callup_nums = spending_data[spending_data['기타'] == 'Call up'].shape[0]
    
    tot_departure_nums = len(df) - spending_count
    FA_departure_nums = income_data[income_data['기타'] == 'FA'].shape[0]
    Loan_departure_nums = income_data[income_data['기타'] == 'Loan'].shape[0]
    
    # 시즌별로 지출과 수입을 저장
    spending_totals.append(tot_spending)
    M_spending_totals.append(tot_spending / 1e6)
    
    income_totals.append(tot_income)
    M_income_totals.append(tot_income / 1e6)
    
    # 소수점 이하 0 제거
    spending_totals = [int(num) if num == int(num) else num for num in spending_totals]
    M_spending_totals = [int(num) if num == int(num) else num for num in M_spending_totals]
    
    income_totals = [int(num) if num == int(num) else num for num in income_totals]
    M_income_totals = [int(num) if num == int(num) else num for num in M_income_totals]
    
    print("-----------------------")
    print("{} 시즌 총 이적료 지출 : {:,.0f} 유로 ({:,}M 유로)\n".format(season_nums[season_count], tot_spending, M_spending_totals[season_count]))
    
    print("총 영입 인원 : {} 명".format(spending_count))
    print("FA 영입 인원 : {} 명".format(FA_arrival_nums))
    print("임대 영입 인원 : {} 명".format(Loan_arrival_nums))
    print("유스 콜업 인원 : {} 명\n".format(Callup_nums))

    print("{} 시즌 총 이적료 수입 : {:,.0f} 유로 ({:,}M 유로)\n".format(season_nums[season_count], tot_income, M_income_totals[season_count]))

    print("총 방출 인원 : {} 명".format(tot_departure_nums))
    print("FA 방출 인원 : {} 명".format(FA_departure_nums))
    print("임대 이적 인원 : {} 명".format(Loan_departure_nums))    
    print("-----------------------")
    
    season_count += 1

print("-----------------------")
print("가장 비싸게 산 선수")
print("{}시즌 {}".format(max_spending_player_season, max_spending_player))
print("금액 : {:,.0f}M 유로".format(max_spending_player_spend))

print("\n가장 비싸게 판 선수")
print("{}시즌 {}".format(max_income_player_season, max_income_player))
print("금액 : {:,.0f}M 유로".format(max_income_player_earn))
print("-----------------------")

# 결과 데이터를 데이터프레임으로
results = {
    "시즌": season_nums,
    "총 이적료 지출(유로)": M_spending_totals,
    "총 이적료 수입(유로)": M_income_totals,
}

df_results = pd.DataFrame(results)

# 넷스펜딩 컬럼 추가
df_results["넷스펜딩(유로)"] = df_results["총 이적료 지출(유로)"] - df_results["총 이적료 수입(유로)"]

# 전체 년도 데이터 행 추가
all_row = {'시즌' : '합계', '총 이적료 지출(유로)' : all_spending, '총 이적료 수입(유로)' : all_income, '넷스펜딩(유로)' : all_net_spending}
df_results = pd.concat([df_results, pd.DataFrame([all_row])], ignore_index=True)

title = '{} ~ {} 시즌'.format(season_nums[0], season_nums[-1])

# 단위 가시성(M) 추가
df_results.iloc[:, 1:] = df_results.iloc[:, 1:].applymap(lambda x: f"{round(x,2)}M" if x.is_integer() else f"{round(x,2)}M")
df_results = pd.DataFrame(df_results)

# '나이' column에 대한 주요 통계 정보
age_summary = all_season_data['나이'].describe()
print(age_summary)

# 결과 데이터프레임을 엑셀 파일로 저장
excel_output_file_path = f"C:/Users/yth21/Desktop/TH/data/Football/results/{title}.xlsx"
df_results.to_excel(excel_output_file_path, index=False)

width = 0.35
x = np.arange(len(season_nums))

fig, ax = plt.subplots(figsize=(12, 6))
rects1 = ax.bar(x - width/2, M_spending_totals, width, label='이적료 지출', color='red')
rects2 = ax.bar(x + width/2, M_income_totals, width, label='이적료 수입', color='blue')

ax.set_xlabel('시즌')
ax.set_ylabel('이적료(유로)')
ax.set_title(f'최근 10시즌 이적료 지출 및 수입 ({title})')
ax.set_xticks(x)
ax.set_xticklabels(season_nums)
ax.legend()

# y 축 레이블 형식 변경 (백만 단위로 표시)
ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: '{:,.0f}M'.format(x)))

plt.xticks(rotation=45)
plt.tight_layout()

plt.savefig(f"C:/Users/yth21/Desktop/TH/data/Football/results/{title}.png", dpi=300)
plt.show()

# 엑셀 파일 로드
wb = openpyxl.load_workbook(excel_output_file_path)

# 활성 시트 선택
ws = wb.active

net_spending_idx = df_results.columns.get_loc('넷스펜딩(유로)') + 1

# 넷스펜딩 셀 가시성 개선
for row in ws.iter_rows(min_row=2, max_col=net_spending_idx, max_row=ws.max_row):
    cell = row[net_spending_idx - 1]
    cell_value = cell.value
    numeric_part = re.search(r'-?\d+', cell_value).group()  # 숫자 부분 추출
    value = int(numeric_part)
    if value < 0:
        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 하늘색
    elif value > 0:
        cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # 붉은색

# 셀 가운데 정렬 및 셀 너비 조절
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
        cell.alignment = Alignment(horizontal='center', vertical='center')
    adjusted_width = (max_length + 2) * 1.5
    ws.column_dimensions[column_letter].width = adjusted_width

# 이미지 삽입
img = Image(f"C:/Users/yth21/Desktop/TH/data/Football/results/{title}.png")
img.width, img.height = 1000, 500

# 이미지를 삽입할 셀 위치 설정 (예: A10)
ws.add_image(img, "F1")

wb.save(excel_output_file_path)
